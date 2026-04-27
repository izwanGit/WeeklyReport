import streamlit as st

# PETRONAS Brand Color Palette
PETRONAS_TEAL       = "#00B1A9"  # Primary
PETRONAS_PURPLE     = "#763F98"
PETRONAS_BLUE       = "#20419A"
PETRONAS_YELLOW     = "#FDB924"
PETRONAS_LIME_GREEN = "#BFD730"

# SVG icon library (inline, no emoji)
_SVG_ICONS = {
    "check": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg>',
    "alert": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"></path><line x1="12" y1="9" x2="12" y2="13"></line><line x1="12" y1="17" x2="12.01" y2="17"></line></svg>',
    "info": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><line x1="12" y1="16" x2="12" y2="12"></line><line x1="12" y1="8" x2="12.01" y2="8"></line></svg>',
    "x-circle": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><line x1="15" y1="9" x2="9" y2="15"></line><line x1="9" y1="9" x2="15" y2="15"></line></svg>',
    "folder": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 19a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h5l2 3h9a2 2 0 0 1 2 2z"></path></svg>',
    "mail": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M4 4h16c1.1 0 2 .9 2 2v12c0 1.1-.9 2-2 2H4c-1.1 0-2-.9-2-2V6c0-1.1.9-2 2-2z"></path><polyline points="22,6 12,13 2,6"></polyline></svg>',
    "shield": '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"></path></svg>',
}

def _get_svg_icon(name: str, color: str = PETRONAS_TEAL) -> str:
    svg = _SVG_ICONS.get(name, "")
    return svg.format(color=color) if svg else ""

def petronas_alert(message: str, type: str = "info", icon: str = "info"):
    # PETRONAS Brand Colors
    colors = {
        "success": (PETRONAS_LIME_GREEN, "rgba(191,215,48,0.15)"),
        "info":    (PETRONAS_TEAL,       "rgba(0,177,169,0.15)"),
        "warning": (PETRONAS_YELLOW,     "rgba(253,185,36,0.15)"),
        "error":   (PETRONAS_PURPLE,     "rgba(118,63,152,0.15)"),
        "blue":    (PETRONAS_BLUE,       "rgba(32,65,154,0.15)"),
    }
    # Map type to default icon
    default_icons = {"success": "check", "info": "info", "warning": "alert", "error": "x-circle", "blue": "info"}
    icon_name = icon if icon in _SVG_ICONS else default_icons.get(type, "info")
    border_color, bg_color = colors.get(type, colors["info"])
    icon_svg = _get_svg_icon(icon_name, border_color)
    icon_html = f"<span style='margin-right: 10px; flex-shrink: 0; display: flex; align-items: center;'>{icon_svg}</span>" if icon_svg else ""
    html = f'''<div style="background-color: {bg_color}; border-left: 3px solid {border_color}; padding: 8px 12px; border-radius: 4px; margin-bottom: 8px; font-family: sans-serif; font-size: 0.82rem; color: #1E293B; display: flex; align-items: center; line-height: 1.4;">{icon_html}<div>{message}</div></div>'''
    st.markdown(html, unsafe_allow_html=True)

import pandas as pd
import json
import os
import datetime
import time
import requests
from jinja2 import Environment, FileSystemLoader
import sys
import tempfile
import base64
import traceback
import io

# Conditional import for win32com
try:
    if sys.platform == 'win32':
        import win32com.client as win32
    else:
        win32 = None
except ImportError:
    win32 = None

# ----------------------------------------------------
# Configuration & Constants
# ----------------------------------------------------
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
    EXE_DIR  = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    EXE_DIR  = BASE_DIR

HISTORY_FILE  = os.path.join(EXE_DIR,  "history.json")
TEMPLATE_FILE = os.path.join(BASE_DIR, "template.html")

# ----------------------------------------------------
# Dashboard Auto-Fetch Configuration
# ----------------------------------------------------
DASHBOARD_URL = (
    "https://mygenieplus-ir1.onbmc.com/dashboards/api/datasources/proxy"
    "/uid/Uf8LY07Vk/api/arsys/v1.0/report/arsqlquery"
)

DASHBOARD_HEADERS = {
    "accept":             "application/json, text/plain, */*",
    "content-type":       "application/json",
    "x-ar-client-type":  "4021",
    "x-ds-authorization": "IMS-JWT JWT PLACEHOLDER",
    "x-grafana-device-id": "c038f697c5ec05209addd40a9fbf77bb",
    "x-grafana-org-id":  "204007533",
    "x-requested-by":    "undefined",
    "origin":  "https://mygenieplus-ir1.onbmc.com",
    "referer": "https://mygenieplus-ir1.onbmc.com/",
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/147.0.0.0 Safari/537.36 Edg/147.0.0.0"
    ),
}

MYGENIE_DOMAIN = "mygenieplus-ir1.onbmc.com"

COOKIE_BRIDGE_URL = "http://localhost:17731"

# ----------------------------------------------------
# Auto Cookie Reader (Cookie Bridge & browser-cookie3)
# ----------------------------------------------------
def parse_raw_cookie(raw_string: str) -> dict:
    """Parses a raw HTTP Cookie header string into a dictionary."""
    cookies = {}
    if not raw_string:
        return cookies
        
    for item in raw_string.split(";"):
        if "=" in item:
            k, v = item.strip().split("=", 1)
            cookies[k] = v
            
    return cookies

def load_cached_cookie():
    try:
        with open("cookie_cache.txt", "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        return ""

def save_cached_cookie(cookie_str):
    try:
        with open("cookie_cache.txt", "w", encoding="utf-8") as f:
            f.write(cookie_str)
    except Exception:
        pass


@st.dialog("Manual Data Sync")
def show_cookie_modal():
    st.markdown("""
    <p style="color: #4A5568; font-size: 0.9rem; margin-bottom: 15px;">Follow these steps to securely pull live data without IT restrictions:</p>
    <div style="background: #F7FAFC; padding: 15px; border-radius: 6px; border: 1px solid #E2E8F0; margin-bottom: 15px;">
        <ol style="margin: 0; padding-left: 20px; color: #2D3748; font-size: 0.85rem; line-height: 1.6;">
            <li>Open Edge and go to the <b>MyGenie Dashboard</b>.</li>
            <li>Press <b>F12</b> to open Developer Tools.</li>
            <li>Go to the <b>Network</b> tab and <b>Refresh</b> the page.</li>
            <li>Click the very first request, scroll down to <b>Request Headers</b>.</li>
            <li>Right-click the <b>Cookie</b> value and select <b>Copy value</b>.</li>
        </ol>
    </div>
    """, unsafe_allow_html=True)
    
    raw_cookie = st.text_area("Paste 'Cookie' String Here:", height=100)
    
    col1, col2 = st.columns(2)
    if col1.button("Sync Data", use_container_width=True, type="primary"):
        if raw_cookie.strip():
            parsed = parse_raw_cookie(raw_cookie)
            wo_res = fetch_open_wo(parsed)
            inc_res = fetch_open_inc(parsed)

            if wo_res is None and inc_res is None:
                st.session_state.sync_status = "Invalid cookie or session expired."
                st.session_state.sync_error = True
            else:
                if wo_res is not None:
                    st.session_state.auto_wo = wo_res
                if inc_res is not None:
                    st.session_state.auto_inc = inc_res

                save_cached_cookie(raw_cookie.strip())
                st.session_state.sync_status = "Data Synced Successfully!"
                st.session_state.sync_error = False
                st.session_state.master_sync_clicked = True
                st.rerun()
        else:
            petronas_alert("Please paste the cookie string first.", type="error")

    if col2.button("Cancel", use_container_width=True):
        st.rerun()


def _year_start_ms() -> int:
    """Unix-ms for Jan 1 of the current year."""
    year = datetime.date.today().year
    return int(time.mktime(time.strptime(f"{year}-01-01", "%Y-%m-%d"))) * 1000


def _now_ms() -> int:
    """Current time in Unix-ms."""
    return int(time.time() * 1000)


def _post_query(sql: str, cookies: dict):
    """
    POST an AR SQL query to the BMC Helix datasource proxy.
    Returns the integer in rows[0][0], or None on any failure.
    """
    payload = {
        "date_format":      "DD/MM/YYYY",
        "date_time_format": "DD/MM/YYYY HH:MM:SS",
        "output_type":      "Table",
        "sql":              sql,
    }
    try:
        resp = requests.post(
            DASHBOARD_URL,
            headers=DASHBOARD_HEADERS,
            cookies=cookies,
            json=payload,
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
        return int(data[0]["rows"][0][0])
    except Exception:
        return None


def fetch_open_wo(cookies: dict):
    """Fetch open Work Order count (year-to-date, MYCAREERX SUPPORT)."""
    from_ms, to_ms = _year_start_ms(), _now_ms()
    sql = f"""SELECT DISTINCT
COUNT(DISTINCT`SRM:Request`.`Request Number`) AS C1

FROM
`SRM:Request`
INNER JOIN `AR System Schema`.`WOI:WorkOrder`
ON ( `SRM:Request`.`Request Number` = `WOI:WorkOrder`.`SRID`  AND `WOI:WorkOrder`.`Work Order ID` LIKE '%ICT_WO%' AND `WOI:WorkOrder`.`Work Order ID` IS NOT NULL)
LEFT OUTER JOIN (`SLM:Measurement` AS `SLM:Measurement1`)
ON (`WOI:WorkOrder`.`Work Order ID` = `SLM:Measurement1`.`ApplicationUserFriendlyID`)
LEFT OUTER JOIN (`SLM:Measurement` AS `SLM:Measurement2`)
ON (`SRM:Request`.`Request Number` = `SLM:Measurement2`.`ApplicationUserFriendlyID`)
WHERE(
`WOI:WorkOrder`.`ASORG` IN ('ARJDBC6460AC66AB204CA7BE8869BB9AF532F9') 
AND `WOI:WorkOrder`.`ASGRP` IN ('MYCAREERX SUPPORT') 
AND `WOI:WorkOrder`.`Request Assignee` IN ('ARJDBC6460AC66AB204CA7BE8869BB9AF532F9')
AND `SRM:Request`.`Submit Date` between {from_ms}/1000 AND {to_ms}/1000
AND `SRM:Request`.`Status` Not IN ('Cancelled', 'Closed')
)
LIMIT 50000 OFFSET 0"""
    return _post_query(sql, cookies)


def fetch_open_inc(cookies: dict):
    """Fetch open Incident count (year-to-date, MYCAREERX SUPPORT)."""
    from_ms, to_ms = _year_start_ms(), _now_ms()
    sql = f"""SELECT DISTINCT
COUNT(DISTINCT`HPD:Help Desk`.`Incident Number`) AS C1
FROM
`HPD:Help Desk`
INNER JOIN (`SLM:Measurement` )
ON (`HPD:Help Desk`.`Incident Number` = `SLM:Measurement`.`ApplicationUserFriendlyID`
AND `HPD:Help Desk`.`Incident Number` LIKE '%ICT_INC%' AND `HPD:Help Desk`.`Incident Number` IS NOT NULL AND `SLM:Measurement`.`SLACategory` = 'Service Level Agreement' 
AND `HPD:Help Desk`.`Status` IN ('Assigned','Pending','In Progress'))
WHERE 
(
`HPD:Help Desk`.`Reported Date` between {from_ms}/1000 AND {to_ms}/1000
AND `HPD:Help Desk`.`Assigned Group` In ('MYCAREERX SUPPORT')
AND `HPD:Help Desk`.`Assigned Support Organization` IN ('ARJDBC6460AC66AB204CA7BE8869BB9AF532F9')
AND `HPD:Help Desk`.`Assignee` IN ('ARJDBC6460AC66AB204CA7BE8869BB9AF532F9')
)
LIMIT 50000 OFFSET 0"""
    return _post_query(sql, cookies)


# ----------------------------------------------------
# Helper Functions
# ----------------------------------------------------
MIN_NUMERIC_ROWS = 1


def find_col(df, target):
    t = target.strip().lower()
    for c in df.columns:
        if str(c).strip().lower() == t:
            return c
    return None


def clean_status(val):
    if pd.isna(val):
        return ""
    return str(val).strip().lower()


def is_active_status(status_str):
    s = clean_status(status_str)
    if not s:
        return False
    inactive = {"closed", "resolved", "cancelled", "canceled"}
    return s not in inactive


def _has_required_columns(df, required_columns):
    sheet_cols = {str(c).strip().lower() for c in df.columns}
    return all(rc.strip().lower() in sheet_cols for rc in required_columns)


def _count_numeric_rows(df, ageing_col_name):
    col = find_col(df, ageing_col_name)
    if col is None:
        return 0
    numeric = pd.to_numeric(df[col], errors='coerce')
    return int(numeric.notna().sum())


def detect_valid_sheet(xl, required_columns, ageing_col_name):
    for sheet_name in xl.sheet_names:
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name)
            if df.empty:
                continue
            if not _has_required_columns(df, required_columns):
                continue
            if _count_numeric_rows(df, ageing_col_name) < MIN_NUMERIC_ROWS:
                continue
            return sheet_name, df
        except Exception:
            continue
    return None, None


def detect_wo_sheet(xl, sr_sheet_name):
    wo_required = {"Service Request Ageing Days", "Work Order ID", "Work Order Status"}
    for sheet_name in xl.sheet_names:
        if sheet_name == sr_sheet_name:
            continue
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name)
            if df.empty:
                continue
            if not _has_required_columns(df, wo_required):
                continue
            if _count_numeric_rows(df, "Service Request Ageing Days") < MIN_NUMERIC_ROWS:
                continue
            return sheet_name, df
        except Exception:
            continue
    return None, None


def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            petronas_alert(f"Error reading history file: {e}", type="error")
            return []
    return []


def save_history(history):
    try:
        with open(HISTORY_FILE, 'w') as f:
            json.dump(history, f, indent=4)
        return True
    except Exception as e:
        petronas_alert(f"Error saving history file: {e}", type="error")
        return False


# Fixed CC list — never changes
OUTLOOK_CC = (
    "yusrinah.mohamed@petronas.com.my; "
    "norhaiza.awang@petronas.com.my; "
    "aisyoul.zainon@petronas.com.my; "
    "prashant.k.singh@oracle.com; "
    "manesh.kallil@oracle.com; "
    "rozaire@petronas.com.my; "
    "jayanthan.sankar@petronas.com.my"
)

CONTACTS_FILE = os.path.join(EXE_DIR, "contacts.json")

def load_contacts() -> dict:
    """Load name→email mapping from contacts.json."""
    try:
        with open(CONTACTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {k: v for k, v in data.items() if not k.startswith("_")}
    except Exception:
        return {}

def resolve_assignee_emails(ticket_lists: list) -> tuple:
    """
    Given a list of ticket dicts, extract unique assignee names,
    look them up in contacts.json, and return (found_emails, missing_names).
    """
    contacts = load_contacts()
    seen = set()
    found, missing = [], []
    for ticket in ticket_lists:
        name = str(ticket.get("Assignee", "")).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        if name in contacts:
            found.append(contacts[name])
        else:
            missing.append(name)
    return found, missing


DEPT_FILTER = "MYCAREERSUPPORT"


def _filter_by_exact_col(ws, col_name: str, dept: str = DEPT_FILTER) -> int:
    """
    Delete rows where the exact column `col_name` does NOT contain `dept`.
    Returns count of remaining data rows.
    """
    headers = [cell.value for cell in ws[1]]
    dept_col = None
    for i, h in enumerate(headers, start=1):
        if h and str(h).strip() == col_name:
            dept_col = i
            break
    if dept_col is None:
        return max(0, ws.max_row - 1)

    to_delete = []
    for row in ws.iter_rows(min_row=2):
        val = str(row[dept_col - 1].value or "").strip().upper()
        if dept.upper() not in val:
            to_delete.append(row[0].row)
    for row_num in reversed(to_delete):
        ws.delete_rows(row_num)
    return max(0, ws.max_row - 1)


def _filter_by_wo_ids(ws, surviving_ids: set) -> None:
    """
    For sheets without the assignee group column, keep only rows
    whose Work Order ID/No. is in `surviving_ids`.
    """
    if not surviving_ids:
        return
    headers = [cell.value for cell in ws[1]]
    wo_id_col = None
    for i, h in enumerate(headers, start=1):
        if h and "work order" in str(h).lower() and (
            "id" in str(h).lower() or "no" in str(h).lower()
        ):
            wo_id_col = i
            break
    if wo_id_col is None:
        return
    to_delete = []
    for row in ws.iter_rows(min_row=2):
        val = str(row[wo_id_col - 1].value or "").strip()
        if val not in surviving_ids:
            to_delete.append(row[0].row)
    for row_num in reversed(to_delete):
        ws.delete_rows(row_num)


def _count_ageing_gt(ws, threshold: int) -> int:
    """
    Count data rows where the ageing-days column value > threshold.
    """
    headers = [cell.value for cell in ws[1]]
    ageing_col = None
    for i, h in enumerate(headers, start=1):
        if h and "ageing" in str(h).lower() and "day" in str(h).lower():
            ageing_col = i
            break
    if ageing_col is None:
        return 0
    count = 0
    for row in ws.iter_rows(min_row=2):
        try:
            val = float(row[ageing_col - 1].value or 0)
            if val > threshold:
                count += 1
        except (TypeError, ValueError):
            pass
    return count


def _update_cover_number(ws, count: int) -> None:
    """
    1. Delete all images/drawings from the cover sheet.
    2. Place the filtered count at E20:J20 with large PETRONAS teal font.
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing

    # ── Remove all drawings & images ─────────────────────────────
    ws._images = []
    try:
        ws._drawing = SpreadsheetDrawing()
    except Exception:
        pass

    # ── Unmerge E20:J20 if already merged ────────────────────────
    for rng in list(ws.merged_cells.ranges):
        r = str(rng)
        if "E20" in r or "F20" in r:
            ws.unmerge_cells(r)

    # ── Write count into E20:J20 ─────────────────────────────────
    ws.merge_cells("E20:J20")
    cell = ws["E20"]
    cell.value     = count
    cell.font      = Font(name="Calibri", bold=True, size=96, color="00B1A9")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[20].height = 100


def _build_update_details_sheet(wb, wo_sheet_name: str, report_date: datetime.date):
    """
    Duplicate the WO Ageing sheet as 'Update Details', keep only
    required columns, filter by MYCAREERSUPPORT via 'Work Order Assignee Group',
    and insert 'Update as of DD Mmm' as the leftmost column.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    KEEP_COLS = {
        "Service Request Ageing Days",
        "Work Order ID",
        "Work Order Summary",
        "Customer Full Name (Service Request)",
        "Work Order Status",
        "Work Order Status Reason",
        "Work Order Assignee",
    }

    wo_ws = wb[wo_sheet_name]

    # Determine which column indices (1-based) to keep
    headers = [cell.value for cell in wo_ws[1]]
    keep_indices = set()
    for i, h in enumerate(headers, start=1):
        if h is None:
            continue
        h_str = str(h)
        if h_str.lower().startswith("status as of") or h_str in KEEP_COLS:
            keep_indices.add(i)

    # Remove old copy if exists
    if "Update Details" in wb.sheetnames:
        del wb["Update Details"]

    # Copy sheet & strip unwanted columns
    new_ws = wb.copy_worksheet(wo_ws)
    new_ws.title = "Update Details"
    max_col = new_ws.max_column
    for col_idx in range(max_col, 0, -1):
        if col_idx not in keep_indices:
            new_ws.delete_cols(col_idx)

    # Filter by MYCAREERSUPPORT using exact SR column name
    _filter_by_exact_col(new_ws, "Work Order Assignee Group")

    # Insert 'Update as of DD Mmm' as column A
    new_ws.insert_cols(1)
    col_header = f"Update as of {report_date.strftime('%d %b')}"
    hdr_cell = new_ws.cell(row=1, column=1)
    hdr_cell.value     = col_header
    hdr_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF")
    hdr_cell.fill      = PatternFill("solid", fgColor="00B1A9")
    hdr_cell.alignment = Alignment(horizontal="center", vertical="center")


def process_sr_wo_workbook(workbook_path: str, report_date: datetime.date):
    """
    Full processing of the SR & WO workbook:
      - Filter raw data sheet by 'Work Order Assignee Group' = MYCAREERSUPPORT
      - Filter sheet 3 (no group col) by surviving WO IDs
      - Update cover sheets: delete images, place filtered count at E20:J20
      - Add 'Update Details' sheet
    """
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    sheet_names = wb.sheetnames

    # Identify the primary WO raw data sheet
    wo_data_name = sheet_names[-1]
    for name in sheet_names:
        if "work order" in name.lower() and "ageing" in name.lower():
            wo_data_name = name

    # ── Step 1: Filter raw data by exact group column ─────────────
    _filter_by_exact_col(wb[wo_data_name], "Work Order Assignee Group")

    # ── Step 2: Collect surviving WO IDs ─────────────────────────
    wo_ws = wb[wo_data_name]
    wo_headers = [cell.value for cell in wo_ws[1]]
    wo_id_col = None
    for i, h in enumerate(wo_headers, start=1):
        if h and "work order" in str(h).lower() and (
            "id" in str(h).lower() or "no" in str(h).lower()
        ):
            wo_id_col = i
            break
    surviving_wo_ids = set()
    if wo_id_col:
        for row in wo_ws.iter_rows(min_row=2):
            val = row[wo_id_col - 1].value
            if val:
                surviving_wo_ids.add(str(val).strip())

    # ── Step 3: Filter remaining data sheets by WO IDs ───────────
    for sname in [s for s in sheet_names[2:] if s != wo_data_name]:
        _filter_by_wo_ids(wb[sname], surviving_wo_ids)

    # ── Step 4: Get counts ────────────────────────────────────────
    total_sr = max(0, wo_ws.max_row - 1)
    sr_gt_30 = _count_ageing_gt(wo_ws, 30)

    # ── Step 5: Update cover sheet numbers (delete image → E20:J20)
    if len(sheet_names) >= 1:
        _update_cover_number(wb[sheet_names[0]], total_sr)
    if len(sheet_names) >= 2:
        _update_cover_number(wb[sheet_names[1]], sr_gt_30)

    # ── Step 6: Build Update Details sheet ───────────────────────
    _build_update_details_sheet(wb, wo_data_name, report_date)

    wb.save(workbook_path)


def process_inc_workbook(workbook_path: str, report_date: datetime.date):
    """
    Full processing of the INC workbook:
      - Filter all data sheets by 'Assignee Group' = MYCAREERSUPPORT
      - Update cover sheets: delete images, place filtered count at E20:J20
    """
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    sheet_names = wb.sheetnames

    # Identify the primary INC data sheet
    inc_data_name = sheet_names[-1]
    for name in sheet_names:
        if "incident" in name.lower() and "raw" in name.lower():
            inc_data_name = name
            break

    # ── Filter all data sheets by exact INC group column ─────────
    for sname in sheet_names[2:]:
        _filter_by_exact_col(wb[sname], "Assignee Group")

    # ── Get counts ────────────────────────────────────────────────
    inc_ws    = wb[inc_data_name]
    total_inc = max(0, inc_ws.max_row - 1)
    inc_gt_30 = _count_ageing_gt(inc_ws, 30)

    # ── Update cover sheet numbers (delete image → E20:J20) ──────
    if len(sheet_names) >= 1:
        _update_cover_number(wb[sheet_names[0]], total_inc)
    if len(sheet_names) >= 2:
        _update_cover_number(wb[sheet_names[1]], inc_gt_30)

    wb.save(workbook_path)


def save_excels_to_onedrive(
    report_date: datetime.date, final_sr_wo_file, final_inc_file
) -> tuple:
    """
    Copy/rename the SR&WO and INC Excel files to OneDrive Weekly Report folder,
    then process both workbooks (filter, cover pages, Update Details sheet).
    Returns (sr_wo_dest_path, inc_dest_path).
    """
    import shutil

    user_profile = os.environ.get("USERPROFILE", "")
    date_folder  = report_date.strftime("%d %B")   # e.g. "27 April"
    date_suffix  = report_date.strftime("%d%b")    # e.g. "27Apr"

    target_dir = os.path.join(
        user_profile, "OneDrive - PETRONAS", "Weekly Report", date_folder
    )
    os.makedirs(target_dir, exist_ok=True)

    sr_wo_dest = os.path.join(
        target_dir,
        f"Service Request & Work Order Ageing Raw Data Preview_{date_suffix}.xlsx",
    )
    inc_dest = os.path.join(
        target_dir,
        f"Incident Ageing Raw Data Preview_{date_suffix}.xlsx",
    )

    # ── Write SR & WO file ────────────────────────────────────────
    if isinstance(final_sr_wo_file, str):
        shutil.copy2(final_sr_wo_file, sr_wo_dest)
    else:
        with open(sr_wo_dest, "wb") as f:
            f.write(final_sr_wo_file.getvalue())
    process_sr_wo_workbook(sr_wo_dest, report_date)

    # ── Write INC file ────────────────────────────────────────────
    if isinstance(final_inc_file, str):
        shutil.copy2(final_inc_file, inc_dest)
    else:
        with open(inc_dest, "wb") as f:
            f.write(final_inc_file.getvalue())
    process_inc_workbook(inc_dest, report_date)

    return sr_wo_dest, inc_dest



def push_to_outlook(html_body, subject="Weekly SR & Incident Update", to_emails=None, cc=OUTLOOK_CC, attachments=None):
    if sys.platform != 'win32' or win32 is None:
        petronas_alert("Outlook integration is only supported on Windows machines with pywin32 installed.", type="error")
        return False
    try:
        import pythoncom
        pythoncom.CoInitialize()
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        mail.HTMLBody = html_body
        if to_emails:
            mail.To = "; ".join(to_emails)
        mail.CC = cc
        if attachments:
            for att_path in attachments:
                if os.path.exists(att_path):
                    mail.Attachments.Add(att_path)
        mail.Display(True)
        return True
    except Exception as e:
        petronas_alert(f"Failed to open Outlook draft: {e}", type="error")
        return False
    finally:
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except:
            pass


# ----------------------------------------------------
# Page Config & Premium Styling
# ----------------------------------------------------
st.set_page_config(
    page_title="Weekly Report | PETRONAS",
    page_icon=os.path.join(BASE_DIR, "PETRONAS_LOGO_SQUARE.png"),
    layout="wide",
    initial_sidebar_state="expanded"
)


def _image_to_data_uri(path, mime_type):
    try:
        with open(os.path.join(BASE_DIR, path), 'rb') as f:
            data = f.read()
        return f"data:{mime_type};base64,{base64.b64encode(data).decode()}"
    except:
        return ""


_logo_square_uri   = _image_to_data_uri("PETRONAS_LOGO_SQUARE.png",      "image/png")
_logo_sidebar_uri  = _image_to_data_uri("PETRONAS_LOGO_HORIZONTAL.svg",  "image/svg+xml")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Inter', sans-serif !important;
    }
    [data-testid="stDecoration"],
    [data-testid="stStatusWidget"],
    [data-testid="stSidebarNav"],
    .stDeployButton { display: none !important; visibility: hidden !important; }
    header[data-testid="stHeader"] { background: transparent !important; border-bottom: none !important; }
    [data-testid="stAppViewContainer"] > .main { transition: none !important; }
    [data-testid="stSidebar"] { animation: none !important; }
    [data-testid="stSidebarNav"],
    [data-testid="stSidebarNavItems"],
    [data-testid="stSidebarNavSeparator"],
    [data-testid="stStatusWidget"] { display: none !important; visibility: hidden !important; }
    header[data-testid="stHeader"] { background: #F8FAFC !important; }
    [data-testid="stSidebar"] { border-right: none !important; }
    .main .block-container { padding-top: 0.5rem !important; max-width: 1400px !important; }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 { color: #00B1A9 !important; font-weight: 700 !important; }
    .stButton > button, .stDownloadButton > button {
        background: linear-gradient(135deg, #00B1A9, #008C86) !important;
        color: white !important;
        border: none !important; border-radius: 10px !important;
        font-weight: 600 !important; transition: all 0.2s ease !important;
        box-shadow: 0 4px 10px rgba(0, 177, 169, 0.3) !important;
        padding: 0.6rem 1.4rem !important;
        display: flex !important; align-items: center !important; justify-content: center !important;
    }
    .stButton > button p::before {
        content: ""; display: inline-block; width: 16px; height: 16px;
        background-image: url("data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='white' stroke-width='2.5' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='23 4 23 10 17 10'/%3E%3Cpath d='M20.49 15a9 9 0 1 1-2.12-9.36L23 10'/%3E%3C/svg%3E");
        background-size: contain; background-repeat: no-repeat; margin-right: 8px; margin-top: 1px;
    }
    .stButton > button p { margin: 0 !important; display: flex !important; align-items: center !important; justify-content: center !important; }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #009C95, #007A75) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 14px rgba(0, 177, 169, 0.4) !important;
        color: white !important;
    }
    .stButton > button:active, .stDownloadButton > button:active {
        transform: translateY(0px) !important;
        box-shadow: 0 2px 5px rgba(0, 177, 169, 0.3) !important;
    }
    [data-testid="stMetric"] {
        background: #FFFFFF !important; border: 1px solid #E2E8F0 !important;
        border-left: 4px solid #00B1A9 !important; border-radius: 12px !important;
        padding: 1.1rem 1.2rem !important; box-shadow: 0 2px 8px rgba(0,0,0,0.04) !important;
    }
    [data-testid="stMetricValue"] { color: #00B1A9 !important; font-weight: 800 !important; font-size: 1.8rem !important; }
    [data-testid="stMetricLabel"] { color: #4A5568 !important; font-weight: 500 !important; }
    .stTabs [data-baseweb="tab"] { font-weight: 500 !important; }
    .stTabs [aria-selected="true"] { color: #00B1A9 !important; font-weight: 700 !important; border-bottom-color: #00B1A9 !important; }
    [data-testid="stFileUploader"] {
        border: 2px dashed rgba(0, 177, 169, 0.4) !important;
        border-radius: 12px !important; padding: 16px 20px !important;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h1,
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2,
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 { margin-bottom: 0px !important; }
    [data-testid="stSidebar"] blockquote { margin-bottom: 0px !important; }
    [data-testid="stSidebar"] .stNumberInput,
    [data-testid="stSidebar"] .stDateInput,
    [data-testid="stSidebar"] .stFileUploader { margin-bottom: -10px !important; }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    .stDeployButton, [data-testid="stDeployButton"], [data-testid="stAppDeployButton"] { display: none !important; }
    .genie-link {
        font-size: 0.85rem; font-weight: 500;
        color: #31333F !important; text-decoration: none !important;
        transition: all 0.2s ease !important; cursor: pointer !important;
    }
    .genie-link:hover { color: #00B1A9 !important; text-decoration: none !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<a href="/" target="_self" style="text-decoration:none;display:inline-flex;align-items:center;gap:8px;font-weight:600;color:#64748B;margin-bottom:16px;transition:color 0.2s ease;">
    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <line x1="19" y1="12" x2="5" y2="12"></line>
        <polyline points="12 19 5 12 12 5"></polyline>
    </svg>
    Back to Hub
</a>
""", unsafe_allow_html=True)

st.markdown(f"""
<style>
.banner-title {{color:#FFFFFF!important;text-transform:uppercase!important;font-weight:800!important;text-shadow:0px 2px 4px rgba(0,0,0,0.3)!important;margin:0!important;line-height:1.1!important;white-space:nowrap;font-size:clamp(1.2rem,3.5vw,1.8rem)!important;letter-spacing:0.1px;}}
.banner-subtitle {{color:#FFFFFF!important;font-weight:400!important;text-shadow:0px 1px 3px rgba(0,0,0,0.2)!important;margin:4px 0 0 0!important;white-space:nowrap;font-size:clamp(0.85rem,2vw,1.0rem)!important;opacity:0.95!important;}}
</style>
<div style="display:flex;align-items:center;gap:24px;padding:22px 32px;background-color:#00B1A9;border-radius:20px;margin-bottom:2rem;box-shadow:0 12px 35px rgba(0,177,169,0.25);overflow:hidden;border:1px solid rgba(255,255,255,0.15);">
<img src="{_logo_square_uri}" style="height:80px;flex-shrink:0;filter:drop-shadow(1px 1px 0 white) drop-shadow(-1px -1px 0 white) drop-shadow(1px -1px 0 white) drop-shadow(-1px 1px 0 white);"/>
<div style="min-width:0;">
<h1 class="banner-title">Weekly SR &amp; Incident Report</h1>
<p class="banner-subtitle">Automate your MyGenie Excel exports into production-ready HTML email reports.</p>
</div>
</div>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# Sidebar
# ----------------------------------------------------
with st.sidebar:
    st.markdown(f"""
<div style="text-align:center;padding:8px 0 20px 0;">
    <a href="/" target="_self" style="display:inline-block;">
        <img src="{_logo_sidebar_uri}" style="height:56px;transition:transform 0.2s;cursor:pointer;"
             onmouseover="this.style.transform='scale(1.05)'"
             onmouseout="this.style.transform='scale(1)'"/>
    </a>
</div>
""", unsafe_allow_html=True)

    # --------------------------------------------------
    # Open Ticket Counts (with optional Auto Sync)
    # --------------------------------------------------
    st.markdown("### Open Ticket Counts")

    if "auto_wo" not in st.session_state:
        st.session_state.auto_wo = None
    if "auto_inc" not in st.session_state:
        st.session_state.auto_inc = None
    if "sync_status" not in st.session_state:
        st.session_state.sync_status = None
    if "sync_error" not in st.session_state:
        st.session_state.sync_error = False

    if "master_sync_clicked" not in st.session_state:
        st.session_state.master_sync_clicked = False

    if st.button("Sync Live Data", use_container_width=True):
        st.session_state.master_sync_clicked = True
        cached_cookie = load_cached_cookie()
        if cached_cookie:
            with st.spinner("Trying cached session..."):
                parsed = parse_raw_cookie(cached_cookie)
                wo_res = fetch_open_wo(parsed)
                inc_res = fetch_open_inc(parsed)
                
                if wo_res is not None or inc_res is not None:
                    if wo_res is not None:
                        st.session_state.auto_wo = wo_res
                    if inc_res is not None:
                        st.session_state.auto_inc = inc_res
                    st.session_state.sync_status = "Data Synced Successfully (Using Cached Session)!"
                    st.session_state.sync_error = False
                else:
                    st.session_state.sync_status = "Cached session expired. Please provide a new cookie."
                    st.session_state.sync_error = True
                    show_cookie_modal()
        else:
            show_cookie_modal()

    if st.session_state.sync_status:
        if st.session_state.sync_error:
            petronas_alert(f"Sync Failed: {st.session_state.sync_status}", type="error")
        else:
            petronas_alert(f"{st.session_state.sync_status}", type="success")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("<a href='https://mygenieplus-ir1.onbmc.com/dashboards/d/aegyhutg26kn4a/f350ff42-68d2-5195-bce1-6a86eeaf6336?orgId=204007533&var-ASORG=All&var-AssignedGroup=MYCAREERX%20SUPPORT&var-assignee=All&var-Status=All' target='_blank' class='genie-link'>Open WO ↗</a>", unsafe_allow_html=True)
        sr_open_wo = st.number_input(
            "Open WO",
            min_value=0,
            value=st.session_state.auto_wo if st.session_state.auto_wo is not None else 1,
            step=1,
            help="Total open Work Order ticket count (e.g. 215)",
            label_visibility="collapsed"
        )
    with c2:
        st.markdown("<a href='https://mygenieplus-ir1.onbmc.com/dashboards/d/beg9bk10a07i8e/39afb7b?orgId=204007533&var-Assigned_Support_Org=All&var-AssignedGroup=MYCAREERX%20SUPPORT&var-Assignee=All&var-SLA=All' target='_blank' class='genie-link'>Open INC ↗</a>", unsafe_allow_html=True)
        inc_open_input = st.number_input(
            "Open INC",
            min_value=0,
            value=st.session_state.auto_inc if st.session_state.auto_inc is not None else 1,
            step=1,
            help="Total open Incident ticket count (e.g. 7)",
            label_visibility="collapsed"
        )

    st.markdown("<div style='margin-top: -30px;'></div>", unsafe_allow_html=True)
    st.markdown("### Report Settings")
    report_date = st.date_input("Report Date", datetime.date.today())
    report_date_str = report_date.strftime("%d %B %Y")
    
    # --------------------------------------------------
    # Data Source (with auto SharePoint detection)
    # --------------------------------------------------
    st.markdown("<div style='margin-top: -25px;'></div>", unsafe_allow_html=True)
    st.markdown("### Data Upload")

    user_profile    = os.environ.get('USERPROFILE', '')
    default_inc_path = os.path.join(
        user_profile, "OneDrive - PETRONAS",
        "SAP HR - HCSM Ticket Monitoring Dashboard",
        "Ticketing Data", "Ageing Incident",
        "Incident Ageing Raw Data (Daily).xlsx",
    )
    default_sr_path = os.path.join(
        user_profile, "OneDrive - PETRONAS",
        "SAP HR - HCSM Ticket Monitoring Dashboard",
        "Ticketing Data", "Ageing Service Request",
        "Service Request Ageing Raw Data [Daily].xlsx",
    )

    sync_active = False
    if st.session_state.master_sync_clicked:
        sync_active = os.path.exists(default_inc_path) and os.path.exists(default_sr_path)

    if sync_active:
        petronas_alert("Local Excel files detected in OneDrive folder.", type="info", icon="folder")
        st.caption("These will be used automatically. Upload below to override.")

    st.markdown("<a href='https://mygenieplus-ir1.onbmc.com/dashboards/d/ce3wv282zk1kwd/service-request-and-work-order-ageing-raw-data?orgId=204007533&var-Ownership=All&var-Assignee_Group=MYCAREERX%20SUPPORT&var-Assigned_Support_Org=All' target='_blank' class='genie-link'>SR & WO Excel ↗</a>", unsafe_allow_html=True)
    uploaded_sr_wo = st.file_uploader("SR & WO Excel", type=['xlsx', 'xls'], key="sr_wo", label_visibility="collapsed")
    
    st.markdown("<a href='https://mygenieplus-ir1.onbmc.com/dashboards/d/ddxo5d7th1gqob/incident-ageing-raw-data?orgId=204007533&var-Assignee_Login=All&var-Assigned_Group=MYCAREERX%20SUPPORT&var-Assigned_Support_Org=All&var-enableOverridesForExcel=true' target='_blank' class='genie-link'>Incident Excel ↗</a>", unsafe_allow_html=True)
    uploaded_inc = st.file_uploader("Incident Excel", type=['xlsx', 'xls'], key="inc", label_visibility="collapsed")

    final_sr_wo_file = uploaded_sr_wo if uploaded_sr_wo else (default_sr_path if sync_active else None)
    final_inc_file   = uploaded_inc   if uploaded_inc   else (default_inc_path if sync_active else None)


# ----------------------------------------------------
# Main Processing Logic
# ----------------------------------------------------
if final_sr_wo_file and final_inc_file:
    try:
        def safe_load_excel(file_or_path, name_label):
            if isinstance(file_or_path, str):
                try:
                    with open(file_or_path, "rb") as f:
                        file_bytes = f.read()
                    return pd.ExcelFile(io.BytesIO(file_bytes))
                except PermissionError:
                    petronas_alert(
                        f"<b>Permission Denied!</b> The {name_label} file is currently locked.<br><br>"
                        f"1. Close Microsoft Excel if you have this file open.<br>"
                        f"2. Ensure OneDrive sync is not paused.<br><br>"
                        f"<b>Locked File:</b> <code>{file_or_path}</code>",
                        type="error"
                    )
                    st.stop()
            else:
                file_or_path.seek(0)
                return pd.ExcelFile(file_or_path)

        xl_sr_wo = safe_load_excel(final_sr_wo_file, "SR & WO")
        xl_inc   = safe_load_excel(final_inc_file,   "Incident")

        sr_required  = {"Service Request Ageing Days", "Service Request ID", "Service Request Status", "Work Order Assignee Group"}
        inc_required = {"Incident Ageing Days", "Incident ID", "Status", "Assignee Group"}

        sr_sheet_name, df_sr_raw = detect_valid_sheet(xl_sr_wo, sr_required, "Service Request Ageing Days")

        wo_sheet_name, df_wo_raw = None, pd.DataFrame()
        if sr_sheet_name is not None:
            wo_required_inline = {"Service Request Ageing Days", "Work Order ID", "Work Order Status"}
            if _has_required_columns(df_sr_raw, wo_required_inline):
                wo_sheet_name = sr_sheet_name
                df_wo_raw     = df_sr_raw.copy()
            else:
                w_name, w_df = detect_wo_sheet(xl_sr_wo, sr_sheet_name)
                if w_name is not None:
                    wo_sheet_name = w_name
                    df_wo_raw     = w_df.copy()

        inc_sheet_name, df_inc_raw = detect_valid_sheet(xl_inc, inc_required, "Incident Ageing Days")

        status_msg  = "**Data Source:** "
        status_msg += "Live SharePoint Sync\n" if isinstance(final_sr_wo_file, str) else "Manual Upload\n"
        status_msg += f"Detected → SR: `{sr_sheet_name}`, WO: `{wo_sheet_name}`, INC: `{inc_sheet_name}`"
        petronas_alert(status_msg, type="info")

        if sr_sheet_name is None:
            petronas_alert("Could not locate a valid Service Request sheet.", type="error")
            st.stop()
        if wo_sheet_name is None:
            petronas_alert("Work Order detail sheet not found — detail tables will be empty.", type="warning")
            df_wo_raw = pd.DataFrame()
        if inc_sheet_name is None:
            petronas_alert("Could not locate a valid Incident sheet.", type="error")
            st.stop()

        # --- SR Metric Calculations ---
        df_sr = df_sr_raw.copy()
        sr_assign_grp_col = find_col(df_sr, "Work Order Assignee Group")
        if sr_assign_grp_col:
            df_sr = df_sr[df_sr[sr_assign_grp_col].astype(str).str.strip().str.upper() == "MYCAREERX SUPPORT"]
        else:
            petronas_alert(f"Filter Skipped: Could not find 'Work Order Assignee Group'. Columns: {', '.join(df_sr.columns.astype(str))}", type="warning")

        sr_ageing_col = find_col(df_sr, "Service Request Ageing Days")
        sr_status_col = find_col(df_sr, "Service Request Status")
        df_sr[sr_ageing_col] = pd.to_numeric(df_sr[sr_ageing_col], errors='coerce')
        df_sr = df_sr.dropna(subset=[sr_ageing_col])
        df_sr = df_sr[df_sr[sr_status_col].apply(is_active_status)]

        sr_total        = len(df_sr)
        sr_gt_30_count  = int((df_sr[sr_ageing_col] > 30).sum())
        sr_15_30_count  = int(((df_sr[sr_ageing_col] >= 15) & (df_sr[sr_ageing_col] <= 30)).sum())
        sr_1_14_count   = int(((df_sr[sr_ageing_col] >= 1)  & (df_sr[sr_ageing_col] <= 14)).sum())
        sr_gt_1_count   = int((df_sr[sr_ageing_col] > 1).sum())

        sr_gt_1_pct  = round((sr_gt_1_count  / sr_open_wo * 100) if sr_open_wo  > 0 else 0)
        sr_gt_30_pct = round((sr_gt_30_count / sr_gt_1_count * 100) if sr_gt_1_count > 0 else 0)

        # --- SR Details (from WO sheet) ---
        df_wo = df_wo_raw.copy() if not df_wo_raw.empty else pd.DataFrame()
        wo_assign_grp_col = find_col(df_wo, "Work Order Assignee Group")
        if wo_assign_grp_col is not None and not df_wo.empty:
            df_wo = df_wo[df_wo[wo_assign_grp_col].astype(str).str.strip().str.upper() == "MYCAREERX SUPPORT"]

        wo_ageing_col  = find_col(df_wo, "Service Request Ageing Days") if not df_wo.empty else None
        wo_status_col  = find_col(df_wo, "Work Order Status")
        wo_id_col      = find_col(df_wo, "Work Order ID")
        wo_summary_col = find_col(df_wo, "Work Order Summary")
        wo_customer_col = find_col(df_wo, "Customer Full Name (Service Request)") or find_col(df_wo, "Customer Full Name")
        wo_reason_col  = find_col(df_wo, "Work Order Status Reason")
        wo_assignee_col = find_col(df_wo, "Work Order Assignee")

        sr_ageing_gt_30_tickets = []
        sr_ageing_15_30_tickets = []

        if wo_ageing_col is not None and not df_wo.empty:
            df_wo[wo_ageing_col] = pd.to_numeric(df_wo[wo_ageing_col], errors='coerce')
            df_wo = df_wo.dropna(subset=[wo_ageing_col])
            if wo_status_col:
                df_wo = df_wo[df_wo[wo_status_col].apply(is_active_status)]

            def _safe(row, col):
                if col is None: return ""
                v = row.get(col, "")
                return "" if pd.isna(v) else str(v)

            def extract_wo_records(subset):
                records = []
                for _, row in subset.iterrows():
                    records.append({
                        'SR Ageing':       int(row[wo_ageing_col]),
                        'Work Order No.':  _safe(row, wo_id_col),
                        'Summary':         _safe(row, wo_summary_col),
                        'User/TSG':        _safe(row, wo_customer_col),
                        'WO Status':       _safe(row, wo_status_col),
                        'WO Status Reason': _safe(row, wo_reason_col),
                        'Assignee':        _safe(row, wo_assignee_col),
                    })
                return records

            sr_ageing_gt_30_tickets = extract_wo_records(df_wo[df_wo[wo_ageing_col] > 30])
            sr_ageing_15_30_tickets = extract_wo_records(df_wo[(df_wo[wo_ageing_col] >= 15) & (df_wo[wo_ageing_col] <= 30)])

        # --- INC Metric Calculations ---
        df_inc = df_inc_raw.copy()
        inc_assign_grp_col = find_col(df_inc, "Assignee Group")
        if inc_assign_grp_col:
            df_inc = df_inc[df_inc[inc_assign_grp_col].astype(str).str.strip().str.upper() == "MYCAREERX SUPPORT"]
        else:
            petronas_alert(f"Filter Skipped: Could not find 'Assignee Group'. Columns: {', '.join(df_inc.columns.astype(str))}", type="warning")

        inc_ageing_col = find_col(df_inc, "Incident Ageing Days")
        inc_status_col = find_col(df_inc, "Status")
        inc_active_col = find_col(df_inc, "Active Incident")

        df_inc[inc_ageing_col] = pd.to_numeric(df_inc[inc_ageing_col], errors='coerce')
        df_inc = df_inc.dropna(subset=[inc_ageing_col])
        if inc_active_col:
            df_inc = df_inc[df_inc[inc_active_col].astype(str).str.strip().str.lower() == "yes"]
        if inc_status_col:
            df_inc = df_inc[df_inc[inc_status_col].apply(is_active_status)]

        inc_total       = len(df_inc)
        inc_gt_90_count = int((df_inc[inc_ageing_col] > 90).sum())
        inc_61_90_count = int(((df_inc[inc_ageing_col] >= 61) & (df_inc[inc_ageing_col] <= 90)).sum())
        inc_31_60_count = int(((df_inc[inc_ageing_col] >= 31) & (df_inc[inc_ageing_col] <= 60)).sum())
        inc_15_30_count = int(((df_inc[inc_ageing_col] >= 15) & (df_inc[inc_ageing_col] <= 30)).sum())
        inc_8_14_count  = int(((df_inc[inc_ageing_col] >= 8)  & (df_inc[inc_ageing_col] <= 14)).sum())
        inc_3_7_count   = int(((df_inc[inc_ageing_col] >= 3)  & (df_inc[inc_ageing_col] <= 7)).sum())
        inc_gt_1_count  = int((df_inc[inc_ageing_col] > 1).sum())

        inc_gt_1_pct = round((inc_gt_1_count / inc_open_input * 100) if inc_open_input > 0 else 0)

        # --- History Tracking ---
        history    = load_history()
        short_date = report_date.strftime("%d-%b-%Y")

        new_record = {
            "date":              short_date,
            "sr_count_gt_30":    sr_gt_30_count,
            "sr_count_15_30":    sr_15_30_count,
            "sr_count_1_14":     sr_1_14_count,
            "inc_count_gt_90":   inc_gt_90_count,
            "inc_count_61_90":   inc_61_90_count,
            "inc_count_31_60":   inc_31_60_count,
            "inc_count_15_30":   inc_15_30_count,
            "inc_count_8_14":    inc_8_14_count,
            "inc_count_3_7":     inc_3_7_count,
        }

        existing_idx = next((i for i, h in enumerate(history) if h.get("date") == short_date), None)

        render_history = list(history)
        if existing_idx is not None:
            render_history[existing_idx] = new_record
        else:
            render_history.append(new_record)

        def parse_date(date_str):
            try:    return datetime.datetime.strptime(date_str, "%d-%b-%Y")
            except: return datetime.datetime.min

        render_history.sort(key=lambda x: parse_date(x.get("date", "")))
        if len(render_history) > 4:
            render_history = render_history[-4:]

        trend_dates    = [h.get("date", "")          for h in render_history]
        sr_trend_gt_30 = [h.get("sr_count_gt_30", 0) for h in render_history]
        sr_trend_15_30 = [h.get("sr_count_15_30", 0) for h in render_history]
        sr_trend_1_14  = [h.get("sr_count_1_14",  0) for h in render_history]
        inc_trend_gt_90 = [h.get("inc_count_gt_90", 0) for h in render_history]
        inc_trend_61_90 = [h.get("inc_count_61_90", 0) for h in render_history]
        inc_trend_31_60 = [h.get("inc_count_31_60", 0) for h in render_history]
        inc_trend_15_30 = [h.get("inc_count_15_30", 0) for h in render_history]
        inc_trend_8_14  = [h.get("inc_count_8_14",  0) for h in render_history]
        inc_trend_3_7   = [h.get("inc_count_3_7",   0) for h in render_history]

        # ========== KPI METRICS CARDS ==========
        st.markdown("<div style='margin-top:-10px;'></div>", unsafe_allow_html=True)
        st.markdown("### Key Metrics Overview")
        st.markdown(
            f"<p style='color:#64748B;margin-top:-15px;'>Report date: "
            f"<b style='color:#00B1A9;'>{report_date_str}</b></p>",
            unsafe_allow_html=True,
        )
        st.markdown("<div style='margin-top:-5px;'></div>", unsafe_allow_html=True)

        m1, m2, m3, m4, m5 = st.columns(5)
        with m1: st.metric("Total SR Tickets (Active)", sr_total)
        with m2: st.metric("SR Ageing > 30d",           sr_gt_30_count)
        with m3: st.metric("SR > 1 day %",              f"{sr_gt_1_pct}%")
        with m4: st.metric("Total INC Tickets (Active)", inc_total)
        with m5: st.metric("INC > 1 day %",             f"{inc_gt_1_pct}%")

        # --- Snapshot Management ---
        st.markdown(" ")
        st.markdown("<p style='color:#64748B;font-size:12px;margin-bottom:5px;font-weight:600;'>SNAPSHOT MANAGEMENT</p>", unsafe_allow_html=True)
        if existing_idx is not None:
            c1, c2 = st.columns([1, 4])
            with c1:
                if st.button("Update Saved Snapshot"):
                    history[existing_idx] = new_record
                    history.sort(key=lambda x: parse_date(x.get("date", "")))
                    if len(history) > 4: history = history[-4:]
                    save_history(history)
                    st.rerun()
            with c2:
                petronas_alert(f"{short_date} is already in History. The table below includes it dynamically.", type="info", icon="check")
        else:
            c1, c2 = st.columns([1, 4])
            with c1:
                if st.button("Save Snapshot to History", type="primary"):
                    history.append(new_record)
                    history.sort(key=lambda x: parse_date(x.get("date", "")))
                    if len(history) > 4: history = history[-4:]
                    save_history(history)
                    st.rerun()
            with c2:
                petronas_alert(f"Not yet saved. Click Save to log {short_date} into history.", type="warning")

        st.markdown("")

        # ========== GENERATE HTML ==========
        env         = Environment(loader=FileSystemLoader(BASE_DIR))
        template    = env.get_template("template.html")
        html_output = template.render(
            report_date=report_date_str,
            sr_open_wo=sr_open_wo,
            sr_total=sr_total,
            sr_gt_1_count=sr_gt_1_count,
            sr_gt_30_count=sr_gt_30_count,
            sr_ageing_more_than_1_day_pct=sr_gt_1_pct,
            sr_ageing_more_than_30_days_pct=sr_gt_30_pct,
            sr_ageing_gt_30_tickets=sr_ageing_gt_30_tickets,
            sr_ageing_15_30_tickets=sr_ageing_15_30_tickets,
            inc_open_input=inc_open_input,
            inc_total=inc_total,
            inc_gt_1_count=inc_gt_1_count,
            inc_ageing_more_than_1_day_pct=inc_gt_1_pct,
            trend_dates=trend_dates,
            sr_trend_gt_30=sr_trend_gt_30,
            sr_trend_15_30=sr_trend_15_30,
            sr_trend_1_14=sr_trend_1_14,
            inc_trend_gt_90=inc_trend_gt_90,
            inc_trend_61_90=inc_trend_61_90,
            inc_trend_31_60=inc_trend_31_60,
            inc_trend_15_30=inc_trend_15_30,
            inc_trend_8_14=inc_trend_8_14,
            inc_trend_3_7=inc_trend_3_7,
        )

        email_subject = (
            f"MyCareerX BAU Support Ticket - Ageing Service Request and Incident "
            f"as {report_date.day} {report_date.strftime('%B')}"
        )

        tab_preview, tab_source, tab_export, tab_history = st.tabs(
            ["Email Preview", "HTML Source", "Export Options", "Manage History"]
        )

        with tab_preview:
            st.markdown(f"""
<div style="background-color:#F8FAFC;border:1px solid #E2E8F0;padding:12px 16px;border-radius:8px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.05);display:flex;align-items:center;gap:12px;">
    <span style="background-color:#00B1A9;color:white;font-weight:700;font-size:0.75rem;padding:4px 8px;border-radius:4px;text-transform:uppercase;letter-spacing:0.5px;">Subject</span>
    <span style="color:#334155;font-size:0.95rem;font-weight:600;">{email_subject}</span>
</div>
""", unsafe_allow_html=True)
            st.markdown("""<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:14px;padding:8px;box-shadow:0 4px 12px rgba(0,0,0,0.04);">""", unsafe_allow_html=True)
            st.components.v1.html(html_output, height=900, scrolling=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with tab_source:
            st.code(html_output, language="html")

        with tab_export:
            st.markdown("### Export Actions")
            subject_copy_html = f"""
<html><head>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
body{{margin:0;padding:0;font-family:'Inter',sans-serif;}}
.container{{display:flex;align-items:center;background:#F8FAFC;border:1px solid #E2E8F0;border-radius:10px;padding:8px 12px;gap:12px;}}
.text{{flex-grow:1;color:#1E293B;font-size:0.95rem;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
button{{background:#00B1A9;color:white;border:none;border-radius:6px;padding:6px 14px;font-size:0.8rem;font-weight:600;cursor:pointer;transition:all 0.2s;flex-shrink:0;}}
button:hover{{background:#008C86;transform:translateY(-1px);}}
#msg{{position:absolute;right:80px;color:#00B1A9;font-size:0.75rem;font-weight:700;display:none;}}
</style></head>
<body>
<div style="color:#4A5568;font-size:0.85rem;font-weight:700;margin-bottom:6px;text-transform:uppercase;letter-spacing:0.5px;">Email Subject</div>
<div class="container">
    <div class="text">{email_subject}</div>
    <span id="msg">COPIED!</span>
    <button onclick="copySubject()">COPY</button>
</div>
<script>
function copySubject(){{
    navigator.clipboard.writeText("{email_subject}").then(()=>{{
        const m=document.getElementById("msg");
        m.style.display="inline";
        setTimeout(()=>m.style.display="none",2000);
    }});
}}
</script>
</body></html>"""
            st.components.v1.html(subject_copy_html, height=85)
            st.markdown("<br>", unsafe_allow_html=True)

            exp1, exp2, exp3 = st.columns(3)
            with exp1:
                st.download_button(
                    label="Download .html",
                    data=html_output,
                    file_name=f"Weekly_Report_{report_date.strftime('%Y%m%d')}.html",
                    mime="text/html",
                    use_container_width=True,
                )
            with exp2:
                copy_btn_html = f"""
<html><head>
<style>
body{{margin:0;padding:0;display:flex;flex-direction:column;align-items:center;font-family:sans-serif;}}
button{{background:linear-gradient(135deg,#00B1A9 0%,#008C86 100%);color:white;border:none;border-radius:8px;font-weight:600;padding:0.42rem 1rem;font-size:0.875rem;cursor:pointer;width:100%;height:40px;transition:all 0.3s ease;}}
button:hover{{filter:brightness(1.1);transform:translateY(-1px);}}
#msg{{color:#00B1A9;font-size:0.8rem;font-weight:600;margin-top:5px;display:none;}}
</style></head>
<body>
<button onclick="copyRichText()">Copy Formatted</button>
<div id="msg">Copied to clipboard!</div>
<div id="source" style="display:none;">{base64.b64encode(html_output.encode('utf-8')).decode('utf-8')}</div>
<script>
function copyRichText(){{
    try{{
        const html=decodeURIComponent(escape(window.atob(document.getElementById("source").innerText)));
        navigator.clipboard.write([new ClipboardItem({{"text/html":new Blob([html],{{type:"text/html"}})}})]).then(()=>{{
            const m=document.getElementById("msg");
            m.style.display="block";
            setTimeout(()=>m.style.display="none",3000);
        }}).catch(e=>console.error(e));
    }}catch(e){{console.error(e);}}
}}
</script>
</body></html>"""
                st.components.v1.html(copy_btn_html, height=42)
            with exp3:
                if sys.platform == 'win32':
                    if st.button("Push to Outlook Draft", use_container_width=True):
                        # 1. Save Excel files to OneDrive and get paths for attachment
                        attachments = []
                        try:
                            sr_wo_path, inc_path = save_excels_to_onedrive(
                                report_date, final_sr_wo_file, final_inc_file
                            )
                            attachments = [sr_wo_path, inc_path]
                            petronas_alert(
                                f"Excel files saved to OneDrive: <code>{os.path.dirname(sr_wo_path)}</code>",
                                type="info", icon="folder"
                            )
                        except Exception as e:
                            petronas_alert(f"Could not save Excel files to OneDrive: {e}", type="warning")

                        # 2. Resolve To: recipients from assignees
                        all_tickets = list(sr_ageing_gt_30_tickets) + list(sr_ageing_15_30_tickets)
                        to_emails, missing = resolve_assignee_emails(all_tickets)
                        st.session_state._missing_contacts = missing

                        if missing:
                            # BLOCK — do not open Outlook until all contacts are mapped
                            petronas_alert(
                                f"<b>Blocked: {len(missing)} assignee(s) have no email mapped.</b><br>"
                                + ", ".join(f"<code>{n}</code>" for n in missing)
                                + "<br>Add their emails in <b>Manage Contacts</b> below, then try again.",
                                type="error"
                            )
                        else:
                            # 3. All contacts resolved — push to Outlook
                            if push_to_outlook(html_output, email_subject, to_emails=to_emails, attachments=attachments):
                                petronas_alert("Draft created in Outlook with Excel attachments. Check To/CC fields before sending.", type="success", icon="mail")
                else:
                    st.button("Outlook (Windows Only)", use_container_width=True, disabled=True)

            # ── Contacts Manager ─────────────────────────────────
            st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
            with st.expander("Manage Contacts", expanded=bool(st.session_state.get('_missing_contacts'))):
                contacts = load_contacts()
                missing_now = st.session_state.get('_missing_contacts', [])

                if missing_now:
                    st.markdown(f"""
                    <div style="border-left:3px solid {PETRONAS_YELLOW}; background:rgba(253,185,36,0.08);
                                padding:8px 14px; border-radius:4px; margin-bottom:12px; font-size:0.82rem; color:#1E293B;">
                        <b>Action needed:</b> {len(missing_now)} assignee(s) from the current report are missing emails.
                        Paste their emails below and save.
                    </div>""", unsafe_allow_html=True)

                    updated = dict(contacts)
                    any_filled = False
                    for name in missing_now:
                        c1, c2 = st.columns([2, 3])
                        c1.markdown(
                            f"<div style='padding:8px 0; font-size:0.83rem; color:#1E293B; font-weight:500;'>"
                            f"{name}</div>", unsafe_allow_html=True
                        )
                        email_val = c2.text_input(
                            label=name, label_visibility="collapsed",
                            placeholder="Paste email address here…",
                            key=f"contact_input_{name}"
                        )
                        if email_val.strip():
                            updated[name] = email_val.strip()
                            any_filled = True

                    if any_filled:
                        if st.button("Save to Contacts", type="primary", use_container_width=False):
                            try:
                                with open(CONTACTS_FILE, "w", encoding="utf-8") as f:
                                    json.dump(updated, f, indent=2, ensure_ascii=False)
                                st.session_state._missing_contacts = []
                                petronas_alert("Contacts saved successfully.", type="success", icon="check")
                                st.rerun()
                            except Exception as e:
                                petronas_alert(f"Could not save contacts: {e}", type="error")

                if contacts:
                    st.markdown(f"""
                    <div style="border-top:1px solid #E2E8F0; margin:12px 0 8px 0;"></div>
                    <p style="font-size:0.78rem; color:#718096; margin-bottom:6px;">
                        <b>Known contacts ({len(contacts)})</b>
                    </p>""", unsafe_allow_html=True)
                    for name, email in contacts.items():
                        ca, cb = st.columns([2, 3])
                        ca.markdown(
                            f"<div style='font-size:0.78rem; color:#1E293B; padding:3px 0;'>{name}</div>",
                            unsafe_allow_html=True
                        )
                        cb.markdown(
                            f"<div style='font-size:0.78rem; color:{PETRONAS_TEAL}; padding:3px 0;'>{email}</div>",
                            unsafe_allow_html=True
                        )
                elif not missing_now:
                    petronas_alert("No contacts saved yet. Push to Outlook Draft first to detect assignee names.", type="info")

        with tab_history:
            st.markdown("### Saved Historical Records")
            saved_data = load_history()
            if not saved_data:
                petronas_alert("No records yet. Process and save a snapshot to see it here.", type="info")
            else:
                for idx, h in enumerate(saved_data):
                    st.markdown("<div style='padding:10px;border:1px solid #E2E8F0;border-radius:8px;margin-bottom:8px;'>", unsafe_allow_html=True)
                    col1, col2 = st.columns([5, 1])
                    with col1:
                        st.markdown(
                            f"**{h.get('date')}** &nbsp;|&nbsp; "
                            f"<span style='color:#718096;'>SR &gt;30d: **{h.get('sr_count_gt_30',0)}**</span> &nbsp;|&nbsp; "
                            f"<span style='color:#718096;'>INC &gt;90d: **{h.get('inc_count_gt_90',0)}**</span>",
                            unsafe_allow_html=True,
                        )
                    with col2:
                        if st.button("Delete", key=f"del_{idx}"):
                            saved_data.pop(idx)
                            save_history(saved_data)
                            st.rerun()
                    st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("---")
                if st.button("Clear All History"):
                    save_history([])
                    st.rerun()

    except Exception as e:
        petronas_alert(f"Error processing the files: {e}", type="error")
        st.code(traceback.format_exc(), language="text")

else:
    st.markdown("""
<div style="text-align:center;padding:70px 40px;background:linear-gradient(180deg,#FFFFFF 0%,#F0FAFA 100%);border:1px solid #E2E8F0;border-top:4px solid #00A19C;border-radius:16px;">
    <h2 style="color:#1A202C!important;font-weight:800!important;margin:0 0 10px 0!important;">Data Feed Required</h2>
    <p style="color:#718096;max-width:480px;margin:0 auto;line-height:1.6;font-size:0.95rem;">
        Waiting for live SharePoint sync connection, or upload the manual <code>.xlsx</code> exports via the sidebar.
    </p>
</div>
""", unsafe_allow_html=True)